import re

from typing import Any, List, Tuple

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from xldigest.database.models import (DatamapItem, Project, ReturnItem,
                                      SeriesItem)
from xldigest.process.cell import Cell
from xldigest.process.cleansers import Cleanser
from xldigest.process.datamap import Datamap
from xldigest.process.exceptions import (
    DuplicateReturnError, ProjectNotFoundError, SeriesItemNotFoundError,
    NonExistantReturnError)

from xldigest.database.setup import set_up_session


class TemplateError(Exception):
    pass


class Digest:
    """
    Initialise a Digest object with a Datamap object. Digest.data is a list
    that is empty upon initialisation.

    To populate it from the Datamap.template, call Digest.read_template().
    To populate it from the Datamap.db_file, call Digest.read_project_data().
    The latter can then be written to Datamap.template with Digest.write()

    To create a Digest object for writing to a template:
        - create a template object with the template you wish to populate.
        e.g bicc_template = BICCTemplate(<path-to-file>)

        - create an 'empty' datamap, based on this template:
        e.g base_datamap = Datamap(bicc_template, <path-to-db-file>)

        - create a Digest object:
        e.g. digest = Digest(base_datamap, <series_item-id>)

        - populate the datamap from the database:
        e.g. digest.data.cell_map_from_database()

        - populate the digest.data with either data from the template
        (if the template is non-writable) or from the database. The Digest
        then acts as the intermediadary between the template and database.

        Data from database:
            digest.read_project_data(<series_item-id>, <project-id>)
            digest.data

        Data from template:
            digest.read_template()
            digest.data
    """

    def __init__(self, dm: Datamap, series_item_id: int,
                 project_id: int) -> None:
        self._datamap = dm
        self._data = []  # type: List[Cell]
        self._existing_series_item_ids = \
            self._get_existing_project_and_series_item_ids()[0]
        self._existing_project_ids = \
            self._get_existing_project_and_series_item_ids()[1]
        self._get_existing_return_project_and_series_item_ids()

        self.series_item_id = series_item_id
        self.project_id = project_id

    @property
    def data(self) -> List[Cell]:
        return self._data

    @property
    def datamap(self) -> Datamap:
        return self._datamap

    @classmethod
    def easy_set_up_session_sqlite(cls, database_file: str):
        """
        For use in repls.
        """
        engine = create_engine("sqlite:///" + database_file)
        Session = sessionmaker(bind=engine)
        return Session()

    def _set_up_session(self):
        return set_up_session(self._datamap.db_file)

    def _get_existing_project_and_series_item_ids(self) -> Tuple[Any, Any]:
        """
        Returns a tuple of lists of series_item_ids and project_ids
        in the database.
        """
        session = self._set_up_session()
        project_ids = session.query(Project.id).first()[0]
        series_item_ids = session.query(SeriesItem.id).first()[0]
        session.close()
        return series_item_ids, project_ids

    def _get_existing_return_project_and_series_item_ids(self) -> None:
        """
        Generate a set containing project_ids in returns, and a set containing
        series_item_ids in returns.
        """
        session = self._set_up_session()
        project_ids_in_returns = session.query(ReturnItem.project_id).all()
        series_item_ids_in_returns = session.query(
            ReturnItem.series_item_id).all()
        self._existing_project_ids_in_returns = {
            id[0]
            for id in project_ids_in_returns
        }
        self._existing_series_item_ids_in_returns = {
            id[0]
            for id in series_item_ids_in_returns
        }
        session.close()

    def _check_series_item_exists_in_db(self) -> None:
        """
        Raise SeriesItemNotFoundError if there is no corresponding series_item_id in
        Digest object.
        """
        session = self._set_up_session()
        series_item_ids = session.query(SeriesItem.id).all()
        series_item_ids = [item[0] for item in series_item_ids]
        if self.series_item_id not in series_item_ids:
            session.close()
            raise SeriesItemNotFoundError('SeriesItem not found in database.')
        session.close()

    def _check_project_exists_in_db(self) -> None:
        """
        Raise ProjectNotFoundError if there is no corresponding project_id in
        Digest object.
        """
        session = self._set_up_session()
        project_ids = session.query(Project.id).all()
        project_ids = [item[0] for item in project_ids]
        if self.project_id not in project_ids:
            session.close()
            raise ProjectNotFoundError('Project not found in database.')
        session.close()

    def read_project_data(self) -> None:
        """
        Reads project data from a database, to create a populated cell map
        in Digest.data.
        """
        self._check_series_item_exists_in_db()
        self._check_project_exists_in_db()
        session = self._set_up_session()
        for cell in self._datamap.cell_map:
            # ONLY ACT ON CELLS THAT HAVE A CELL_REFERENCE
            if cell.cell_reference:
                cell.cell_value = session.query(ReturnItem.value).filter(
                    ReturnItem.project_id == self.project_id).filter(
                        ReturnItem.datamap_item_id == DatamapItem.id).filter(
                            ReturnItem.series_item_id ==
                            self.series_item_id).filter(
                                DatamapItem.id == cell.datamap_id[0]).first()
                self.data.append(cell)
        session.close()

    def _generate_file_name_from_return_data(self,
                                             series_item_id: int,
                                             project_id: int) -> str:
        session = self._set_up_session()
        q_str = session.query(SeriesItem.name).filter(
            SeriesItem.id == series_item_id).first()[0]
        proj_str = session.query(Project.name).filter(
            Project.id == project_id).first()[0]
        white_sp = re.compile(r'[\W/]')
        lc_fn = white_sp.sub('_', (proj_str + '_' + q_str))
        return lc_fn

    def write_to_template(self, dir: str) -> str:
        """
        If self._datamap.template is a blank template, then write_to_template()
        will write the datamap.cell_map to it.
        """
        if self._check_for_existing_return():
            if self._datamap.template.writable is False:
                self.read_project_data()
                blank = load_workbook(self._datamap.template.file_name)
                for celldata in self.data:
                    try:
                        blank[celldata.template_sheet][
                            celldata.
                            cell_reference].value = celldata.cell_value[0]
                    except TypeError:
                        blank[celldata.template_sheet][
                            celldata.cell_reference].value = ""
                output_path = (
                    dir + '/' + self._generate_file_name_from_return_data(
                        self.series_item_id, self.project_id) + '.xlsx')
                blank.save(output_path)
                return output_path

            else:
                raise TemplateError(
                    "Cannot write to template which contains source data.")
        else:
            raise NonExistantReturnError(
                "A return with project_id {} and series_item_id {} is not in the database.".
                format(self.project_id, self.series_item_id))

    def _check_for_existing_return(self) -> bool:
        """
        Checks for existence of any returns in database matching qtr_id and
        pjt_id. If there are already records in there with these values,
        we're going to be duplicating returns, therefore this is not allowed.
        """
        if self.project_id in self._existing_project_ids_in_returns \
                and self.series_item_id in self._existing_series_item_ids_in_returns:
            return True
        else:
            return False

    def write_to_database(self) -> None:
        """
        Checks whether there is a series_item and a project in the database that
        corresponds with the Digest data we wish to import into the database.
        If there isn't, an appropriate exception is raised.

        If we go ahead, we then check whether there is already a return in the
        database with the same series_item/project_id combination. If there is,
        we should not be importing this Digest and an appropriate exception
        is raised.

        Otherwise, write the Digest data to the database.
        """
        self._check_series_item_exists_in_db()
        self._check_project_exists_in_db()
        if not self._check_for_existing_return():
            session = self._set_up_session()
            for cell in self._data:
                return_item = ReturnItem(
                    project_id=self.project_id,
                    series_item_id=self.series_item_id,
                    datamap_item_id=cell.datamap_id[0],
                    value=cell.cell_value)
                session.add(return_item)
            session.commit()
            session.close()
        else:
            raise DuplicateReturnError(
                "Existing records in database with series_item_id:"
                " {} project_id: {}".format(self.series_item_id,
                                            self.project_id))

    def read_template(self) -> None:
        """
        Read the relevant values from the template, based on the Datamap.
        Made available in self.data.
        """
        wb = load_workbook(self._datamap.template.file_name)
        for cell in self._datamap.cell_map:
            if cell.cell_reference:
                cell.cell_value = wb[cell.template_sheet][
                    cell.cell_reference].value
                # as long as that value is not None, we cleanse the value
                if cell.cell_value is not None:
                    cleansed = Cleanser(cell.cell_value)
                    cleansed.clean()
                    cell.cell_value = cleansed.string
                self.data.append(cell)
