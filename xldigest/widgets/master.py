"""
A Qt version of the old bcompiler master spreadsheet. Re-written for the new
age...
"""
from typing import Optional, List

from PyQt5 import QtWidgets, QtCore, QtGui
from openpyxl import Workbook

from xldigest import session
from xldigest.database.base_queries import (
    datamap_items_in_return,
    formulate_data_for_master_model,
    project_ids_in_returns_with_series_item_of,
    create_master_friendly_header,
    series_item_ids_in_returns
)
from xldigest.database.models import SeriesItem, ReturnItem, DatamapItem
from xldigest.process.exceptions import NoDataToCreateMasterError


class DatamapCellItem:
    def __init__(self, data: str, x: int, y: int, header=False, imported_session=None) -> None:
        if imported_session:
            self.session = imported_session
        else:
            self.session = session
        self.data = data
        self.header = header
        self.x = x
        self.y = y

    def __repr__(self):
        return "DatamapCellItem(\"{}\")".format(self.data)


class DatamapView:
    returns_added = 0

    def __init__(self, series_item_id: int, imported_session=None) -> None:
        if imported_session:
            self.session = imported_session
        else:
            self.session = session
        self._series_item_id = series_item_id
        self.series_item_name = self._series_item_name()
        self.matrix: List[DatamapCellItem] = []
        self._setup_base_matrix()

    def _setup_base_matrix(self) -> None:
        """
        The matrix will need two base columns: DatamapItem ids (header: "DMI") and
        ReturnItem keys (header: "Key"). Here we set these up.
        :return:
        """
        dmis = self.session.query(DatamapItem.id).all()
        dm_keys = self.session.query(DatamapItem.key).all()
        dmis = [item[0] for item in dmis]
        dm_keys = [item[0] for item in dm_keys]
        self.matrix.append(DatamapCellItem("DMI", 0, 0, header=True))
        self.matrix.append(DatamapCellItem("Key", 1, 0, header=True))
        count = 1  # start adding DMIs from second row down, after header
        for dmi in dmis:
            self.matrix.append(DatamapCellItem(dmi, 0, count))
            count += 1
        count = 1  # start adding DM keys from second row down, after header
        for dm_key in dm_keys:
            self.matrix.append(DatamapCellItem(dm_key, 1, count))
            count += 1

    def _series_item_name(self) -> str:
        name = self.session.query(SeriesItem.name).filter(
            SeriesItem.id == self._series_item_id).first()[0]
        return name

    def add_single_return(self, project_id: int) -> None:
        """
        Take a project_id and its return data to the matrix.
        :param project_id:
        :return: None
        """
        return_data = self.session.query(ReturnItem.value).filter(
            ReturnItem.series_item_id == self._series_item_id,
            ReturnItem.project_id == project_id).all()
        return_data = [item[0] for item in return_data]
        count = 1  # start adding return data from second row down, after header
        for d in return_data:
            self.matrix.append(DatamapCellItem(d, DatamapView.returns_added + 2, count))
            count += 1
        DatamapView.returns_added += 1

    def cell_data(self, x, y) -> Optional[DatamapCellItem]:
        """
        Return DatamapCellItem object at x, y in DatamapView matrix.
        :param x:
        :param y:
        :return: DatamapCellItem
        """
        try:
            gen = (item for item in self.matrix if item.x == x and item.y == y)
            return next(gen)
        except StopIteration:
            return None

    def __str__(self):
        return 'DatamapView for SeriesItem {}'.format(self.series_item_name)

    def __repr__(self):
        return "DatamapView({})".format(self._series_item_id)


class MasterTableModel(QtCore.QAbstractTableModel):
    def __init__(self, data_in, parent=None):
        super().__init__(parent)
        self.data_in = data_in
        self.p_names_on_pop_form = list(self.data_in[0])
        self.headers = create_master_friendly_header(
            self.p_names_on_pop_form, 1)
        self.headers.insert(0, "DMI")
        self.headers.insert(1, "Key")

    def rowCount(self, parent=QtCore.QModelIndex()):
        return len(self.data_in)

    def columnCount(self, parent=QtCore.QModelIndex()):
        return len(self.data_in[0])

    def data(self, index, role):
        if index.isValid() and role == QtCore.Qt.DisplayRole:
            row = index.row()
            col = index.column()
            value = self.data_in[row][col]
            return value

        if index.isValid() and role == QtCore.Qt.BackgroundRole:
            row = index.row()
            col = index.column()
            value = self.data_in[row][col]
            if isinstance(value, (int, float)) and col > 0:
                if value > 10:
                    return QtGui.QColor(23, 243, 24)
            if isinstance(value, str) and col > 0:
                try:
                    value = int(value)
                    if value > 10:
                        return QtGui.QColor(32, 231, 100)
                except:
                    try:
                        value = float(value)
                        if value > 10:
                            return QtGui.QColor(32, 231, 100)
                    except:
                        pass
            else:
                return None

    def headerData(self, section, orientation, role):
        headers = self.headers
        if role == QtCore.Qt.DisplayRole:
            if orientation == QtCore.Qt.Horizontal:
                return headers[section]
            else:
                return section + 1

    def flags(self, index):
        return QtCore.Qt.ItemIsEnabled


class MasterWidget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.series_combo = QtWidgets.QComboBox(self)
        self.set_series_item_combo()
        self.selected_series_item = self.series_combo.itemData(0, QtCore.Qt.UserRole)
        self.series_combo.currentIndexChanged.connect(self._swap_table_slot)

        project_ids = project_ids_in_returns_with_series_item_of(
            self.selected_series_item)  # TODO to get which series_item
        # FIXME - this shit is hard-coded
        try:
            self.datamap_keys = datamap_items_in_return(1, 1)  # TODO likewise - fix hard-cde
        except NoDataToCreateMasterError:
            no_master_diag = QtWidgets.QDialog(self)
            if no_master_diag.exec_():
                print("Done")

        self.table_data = formulate_data_for_master_model(
            self.selected_series_item, project_ids, self.datamap_keys)

        self.tv = QtWidgets.QTableView()
        self.tv.verticalHeader().hide()
        self.proxyModel = QtCore.QSortFilterProxyModel()
        self.tableModel = MasterTableModel(self.table_data)
        self.tv.setModel(self.proxyModel)
        self.proxyModel.setSourceModel(self.tableModel)
        self.tv.horizontalHeader().setStretchLastSection(False)
        col_count = self.tableModel.columnCount()
        self.tv.setColumnWidth(0, 50)
        self.tv.setColumnWidth(1, 300)
        for c in range(2, col_count):
            self.tv.setColumnWidth(c, 200)
        self.sortCaseSensitivityCheckBox = QtWidgets.QCheckBox(
            "Case sensitive sorting")
        self.filterCaseSensitivityCheckBox = QtWidgets.QCheckBox(
            "Case sensitive filter")
        self.filterPatternLineEdit = QtWidgets.QLineEdit()
        self.filterPatternLabel = QtWidgets.QLabel("Filter pattern")
        self.filterPatternLabel.setBuddy(self.filterPatternLineEdit)
        self.filterSyntaxCombo = QtWidgets.QComboBox()
        self.filterSyntaxCombo.addItem("Regular Expression",
                                       QtCore.QRegExp.RegExp)
        self.filterSyntaxCombo.addItem("Wildcard", QtCore.QRegExp.Wildcard)
        self.filterSyntaxCombo.addItem("Fixed string",
                                       QtCore.QRegExp.FixedString)
        self.filterSyntaxLabel = QtWidgets.QLabel("Filter syntax:")
        self.filterSyntaxLabel.setBuddy(self.filterSyntaxCombo)
        self.filterColumnCombo = QtWidgets.QComboBox()
        self.filterColumnCombo.addItem("Key")
        self.filterColumnLabel = QtWidgets.QLabel("Filter column:")
        self.filterColumnLabel.setBuddy(self.filterColumnCombo)

        self.filterPatternLineEdit.textChanged.connect(self.filterRegExChanged)
        self.filterSyntaxCombo.currentIndexChanged.connect(
            self.filterRegExChanged)
        self.filterColumnCombo.currentIndexChanged.connect(
            self.filterColumnChanged)
        self.filterCaseSensitivityCheckBox.toggled.connect(self.sortChanged)

        self.tv.setSortingEnabled(True)

        self.export_button = QtWidgets.QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_master_to_excel_slot)

        proxyGroupBox = QtWidgets.QGroupBox("Master Data")
        proxyLayout = QtWidgets.QGridLayout()
        proxyLayout.addWidget(self.series_combo, 0, 0, 1, 1)
        proxyLayout.addWidget(self.tv, 1, 0, 1, 3)
        proxyLayout.addWidget(self.filterPatternLabel, 2, 0)
        proxyLayout.addWidget(self.filterPatternLineEdit, 2, 1, 1, 2)
        proxyLayout.addWidget(self.filterSyntaxLabel, 3, 0)
        proxyLayout.addWidget(self.filterSyntaxCombo, 3, 1, 1, 2)
        proxyLayout.addWidget(self.filterColumnLabel, 4, 0)
        proxyLayout.addWidget(self.filterColumnCombo, 4, 1, 1, 2)
        proxyLayout.addWidget(self.filterCaseSensitivityCheckBox, 5, 0, 1, 2)
        proxyLayout.addWidget(self.sortCaseSensitivityCheckBox, 5, 2)
        proxyLayout.addWidget(self.export_button, 6, 1)
        proxyGroupBox.setLayout(proxyLayout)

        mainLayout = QtWidgets.QVBoxLayout()

        mainLayout.addWidget(proxyGroupBox)
        self.setLayout(mainLayout)

        self.setWindowTitle("xldigest Master View")

        self.tv.sortByColumn(0, QtCore.Qt.AscendingOrder)
        self.filterColumnCombo.setCurrentIndex(1)

        self.filterPatternLineEdit.setText("")
        self.filterCaseSensitivityCheckBox.setChecked(False)
        self.sortCaseSensitivityCheckBox.setChecked(False)

    def export_master_to_excel_slot(self) -> None:
        """
        Exports the model in the MasterWidget to a new Excel file.
        :return:
        """
        wb = Workbook()
        f_selected = QtWidgets.QFileDialog()
        f_selected.setNameFilter("Excel (*.xlsx)")
        f_selected.setAcceptMode(QtWidgets.QFileDialog.AcceptSave)
        if f_selected.exec_():
            dest_file = "".join([f_selected.selectedFiles()[0], ".xlsx"])
        ws = wb.active
        ws.title = "Master Output"
        capture = []
        rows = self.proxyModel.rowCount()
        cols = self.proxyModel.columnCount()

        for row in range(0, rows):
            line = []
            for col in range(cols):
                i = self.proxyModel.index(row, col)
                line.append(self.proxyModel.data(i, QtCore.Qt.DisplayRole))
            capture.append(tuple(line))

        for h in range(1, cols + 1):
            l = self.proxyModel.headerData(h - 1, QtCore.Qt.Horizontal, QtCore.Qt.DisplayRole)
            ws.cell(column=h, row=1, value=l)

        for line in list(enumerate(capture, 2)):
            c = 1
            for item in line[1]:
                ws.cell(column=c, row=line[0], value=line[1][line[1].index(item)])
                c += 1
        wb.save(dest_file)
        conf_diag = QtWidgets.QDialog(self)
        message_label = QtWidgets.QLabel("File {} created.".format(dest_file))
        button_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok)
        grid = QtWidgets.QGridLayout()
        grid.addWidget(message_label, 0, 0)
        grid.addWidget(button_box, 1, 0)
        conf_diag.setLayout(grid)
        button_box.accepted.connect(conf_diag.accept)
        conf_diag.show()

    def _swap_table_slot(self, index: QtCore.QModelIndex) -> None:
        si = self.series_combo.itemData(index, QtCore.Qt.UserRole)
        project_ids = project_ids_in_returns_with_series_item_of(
            si)
        self.table_data = formulate_data_for_master_model(
            si, project_ids, self.datamap_keys)
        self.tableModel = MasterTableModel(self.table_data, self)
        self.proxyModel.setSourceModel(self.tableModel)
        self.tv.setModel(self.proxyModel)

    def set_series_item_combo(self):
        for item in series_item_ids_in_returns():
            self.series_combo.addItem(item[1], item[0])

    def filterRegExChanged(self):
        syntax = QtCore.QRegExp.PatternSyntax(
            self.filterSyntaxCombo.itemData(
                self.filterSyntaxCombo.currentIndex()))
        caseSensitivity = self.filterCaseSensitivityCheckBox.isChecked()
        regex = QtCore.QRegExp(self.filterPatternLineEdit.text(),
                               caseSensitivity, syntax)
        self.proxyModel.setFilterRegExp(regex)

    def filterColumnChanged(self):
        self.proxyModel.setFilterKeyColumn(
            self.filterColumnCombo.currentIndex() + 1)

    def sortChanged(self):
        self.proxyModel.setSortCaseSensitivity(
            self.sortCaseSensitivityCheckBox.isChecked())
