import csv
import os
from datetime import datetime
import pytest

from openpyxl import Workbook

ws_summary_B5_rand = [
    'Cookfield, Rebuild',
    'Smithson Glenn Park Editing',
    'Brinkles Bypass Havensmere',
    'Folicles On Fire Ltd Extradition',
    'Puddlestein Havelock Underpass',
]

ws_summary_B8_rand = [
    'Aerobics, Maritime and Commerce',
    'TSAD',
    'Special Transport Needs for the Northern Populace',
    'Parenting, Levels and Financial Irregularity',
    'HR',
]


ws_finance_C6_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]

ws_finance_C11_rand = [
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
]


ws_finance_B19_rand = [
    '2012',
    '2013',
    '2011',
    '2018',
    '2002',
    '2007',
]

ws_finance_C18_rand = [
    'Real',
    'Nominal',
]

ws_finance_C36_rand = [
    '2.00',
    '4.20',
    '1.13',
    '12.09',
    '222.07',
]

ws_finance_C44_rand = [
    '12.00',
    '41.20',
    '13.13',
    '122.09',
    '22.07',
]

ws_finance_C77_rand = [
    '29.00',
    '49.23',
    '23.43',
    '1.89',
    '290.37',
]


ws_resources_C7_rand = [
    '9.00',
    '19.00',
    '29.5',
    '12.00',
    '20.5',
]

ws_resources_G17_rand = [
    '9.90',
    '19.22',
    '29.93',
    '1202.89',
    '20.37',
]


ws_resources_I30_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]


ws_resources_J30_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]


ws_resources_J38_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]

ws_approval_C10_rand = [
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
]


ws_approval_F19_rand = [
    'A load of absolute\n horseradish.',
    'When people speak of these kind of things, they are often surprised.',
    'It is very bad here. Completely unacceptable when you think about it.',
    'Never worry too much about it - it wont last forever',
    'There is a forester on this project who is disrupting everything.'
]


ws_approval_B39_rand = [
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
]


ws_assurance_C4_rand = [
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
]


ws_assurance_D10_rand = [
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
    '02-2-2011',
]


ws_resources_E17_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]


@pytest.fixture
def bicc_return():
    wb = Workbook()
    wb.create_sheet('Summary')
    wb.create_sheet('Finance & Benefits')
    wb.create_sheet('Approval & Project milestones')
    wb.create_sheet('Resources')
    wb.create_sheet('Assurance planning')
    wb.create_sheet('GMPP info')

    # Summary fixture
    ws_summary = wb['Summary']
    ws_summary['A5'].value = 'Project/Programme Name'
    ws_summary['B5'].value = ws_summary_B5_rand[0]
    ws_summary['A8'].value = 'DfT Group'
    ws_summary['B8'].value = ws_summary_B8_rand[0]

    # Finance & Benefits fixture
    ws_finance = wb['Finance & Benefits']
    ws_finance['A6'].value = 'SRO Finance Confidence'
    ws_finance['C6'].value = ws_finance_C6_rand[0]
    ws_finance['B11'].value = 'Date of Business Case'
    ws_finance['C11'].value = ws_finance_C11_rand[0]
    ws_finance['A19'].value = 'Index Year'
    ws_finance['B19'].value = ws_finance_B19_rand[0]
    ws_finance['A18'].value = 'Real or Nominal'
    ws_finance['C18'].value = ws_finance_C18_rand[0]
    ws_finance['A36'].value = '2019/2020'
    ws_finance['C36'].value = ws_finance_C36_rand[0]
    ws_finance['A44'].value = 'Total'
    ws_finance['C44'].value = ws_finance_C44_rand[0]
    ws_finance['A77'].value = 'Total WLC (RDEL)'
    ws_finance['C77'].value = ws_finance_C77_rand[0]

    # Resources fixture
    ws_resources = wb['Resources']
    ws_resources['A7'].value = 'SCS(PB2)'
    ws_resources['C7'].value = ws_resources_C7_rand[0]
    ws_resources['A17'].value = 'Total'
    ws_resources['G17'].value = ws_resources_G17_rand[0]
    ws_resources['A30'].value = 'Change Implementation'
    ws_resources['I30'].value = ws_resources_I30_rand[0]
    ws_resources['J30'].value = ws_resources_I30_rand[1]
    ws_resources['G38'].value = 'Overall Assessment'
    ws_resources['J38'].value = ws_resources_J38_rand[0]

    # Approval and Project Milestones fixture
    ws_approvals = wb['Approval & Project milestones']
    ws_approvals['A10'].value = 'SOBC - HMT Approval'
    ws_approvals['C10'].value = ws_approval_C10_rand[0]
    ws_approvals['A19'].value = 'FBC - HMT Approval'
    ws_approvals['F19'].value = ws_approval_F19_rand[0]
    ws_approvals['A39'].value = 'Completion of Construction'
    ws_approvals['B39'].value = ws_approval_B39_rand[0]

    # Assurance fixture
    ws_assurance = wb['Assurance planning']
    ws_assurance['B4'].value = 'Date Created'
    ws_assurance['C4'].value = ws_assurance_C4_rand[0]
    ws_assurance['A10'].value = 'Gate 0 (Programme)'
    ws_assurance['D10'].value = ws_assurance_D10_rand[0]
    ws_assurance['A17'].value = 'Review Point 4 MPRG'
    ws_assurance['E17'].value = 'Amber/Green'

    wb.save('/tmp/test-bicc-return.xlsx')
    yield '/tmp/test-bicc-return.xlsx'
    os.unlink('/tmp/test-bicc-return.xlsx')


@pytest.fixture
def mock_datamap_source_file():
    data = [
        ['Project/Programme Name', 'Summary', 'B5', 'red', 'white', '', 'Yes/No'],
        ['SRO Sign-Off', 'Summary', 'B49', 'red', 'white', '', 'Yes/No'],
        ['GMPP - FD Sign-Off', 'Summary'],
        ['GMPP - Person completing this return'],
        ['GMPP - Single Point of Contact Email Address'],
        ['GMPP - Single Point of Contact (SPOC)'],
        ['GMPP - Email Address'],
        ['Reporting period (GMPP - Snapshot Date)', 'Summary', 'G3', 'red', 'white', '', 'Yes/No'],
        ['Quarter Joined', 'Summary', 'I3', 'red', 'white', '', 'Yes/No'],
        ['GMPP - Sub-portfolio'],
        ['Index Year', 'Finance & Benefits', 'B19', 'red', 'white', '', 'Yes/No'],
        ['Real or Nominal - Baseline', 'Finance & Benefits', 'C18', 'red', 'white', '', 'Yes/No'],
        ['GMPP/quarter formally joined'],
        ['GMPP (GMPP – formally joined GMPP)', 'Summary', 'G5', 'red', 'white', '', 'Yes/No'],
        ['IUK top 40', 'Summary', 'G6', 'red', 'white', '', 'Yes/No'],
        ['Top 37', 'Summary', 'I5', 'red', 'white', '', 'Yes/No'],
        ['DfT Business Plan', 'Summary', 'I6', 'red', 'white', '', 'Yes/No'],
        ['GMPP - IPA ID Number', 'Summary', 'C6', 'red', 'white', '', 'Yes/No'],
        ['DFT ID Number', 'Summary', 'B6', 'red', 'white', '', 'Yes/No'],
        ['Working Contact Name', 'Summary', 'H8', 'red', 'white', '', 'Yes/No'],
        ['Working Contact Telephone', 'Summary', 'H9', 'red', '', ''],
        ['Working Contact Email', 'Summary', 'H10', 'red', 'white', '', 'Yes/No'],
        ['DfT Group', 'Summary', 'B8', 'red', 'yellow', '', 'DfT Group'],
        ['Significant Steel Requirement', 'Finance & Benefits', 'D15', 'blue', 'yello', '', 'Yes/No'],
        ['SRO Finance confidence', 'Finance & Benefits', 'C6', 'green', 'red', '',  'RAG_Short'],
        ['BICC approval point', 'Finance & Benefits', 'E9', 'orange', 'red', '', 'Business Cases'],
        ['Assurance MM2 Latest Approved Baseline', 'Assurance planning', 'C10', 'red', 'white', '', 'Yes/No'],
        ['Approval MM11 Notes','Approval & Project milestones', 'F19', 'red', 'yellow', '', 'Yes/No'],
        ['SCS PB2 No public sector', 'Resources', 'C7', 'red', 'white', '', 'Yes/No'],
        ['Project MM31 Original Baseline', 'Approval & Project milestones', 'B39', 'red', 'white', 'd/mm/yy', 'Yes/No'],
        ['Change Implementation - Now', 'Resources', 'I30', 'black', 'yellow', 'd/mm/yy', 'Capability RAG']
    ]
    with open('/tmp/mock_datamap.csv', 'w') as f:
        datamap_writer = csv.writer(f, delimiter=',')
        f.write('cell_key,template_sheet,cell_reference,bg_colour,fg_colour'
                ',number_format,verification_list\n')
        for item in data:
            datamap_writer.writerow(item)
    yield '/tmp/mock_datamap.csv'
    os.unlink('/tmp/mock_datamap.csv')
